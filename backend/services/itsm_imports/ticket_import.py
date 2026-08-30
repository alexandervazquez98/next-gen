"""Atomic XLSX ticket import — WU 7.

Imports a ticket workbook with three reference sheets (``Ref - Incident
Services``, ``Ref - Service Request Services``, ``Ref - Active Users``).
The import path is all-or-nothing: it acquires per-user PostgreSQL advisory
locks for every distinct assignee in sorted order, holds them through the
single Neo4j write transaction, and releases them whether the write commits
or rolls back.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Protocol

from openpyxl import Workbook
from pydantic import ValidationError

from models.itsm import TicketFolioCreate
from services.user_lock import acquire_user_locks_in_order
from .errors import IMPORT_VALIDATION_FAILED, ImportValidationError
from .workbook import (
    DEFAULT_MAX_SIZE_BYTES,
    guard_xlsx_payload,
    open_workbook,
    read_data_rows,
    read_header_row,
    validate_required_headers,
)


# ---------------------------------------------------------------------------
# Template contract
# ---------------------------------------------------------------------------

TICKET_SHEET = "Ticket Import"
TICKET_REF_INCIDENT = "Ref - Incident Services"
TICKET_REF_SERVICE_REQUEST = "Ref - Service Request Services"
TICKET_REF_USERS = "Ref - Active Users"

TICKET_REQUIRED_HEADERS = (
    "type",
    "title",
    "description",
    "service_catalog_id",
    "assignee_username",
)


# ---------------------------------------------------------------------------
# Repository seams — kept narrow so tests can stub them easily.
# ---------------------------------------------------------------------------


class CatalogRepository(Protocol):
    def get_by_id(self, service_id: str) -> dict | None: ...
    def list(self, limit: int = 100) -> list[dict]: ...


class UserRepository(Protocol):
    def get_by_username(self, db: Any, username: str) -> Any: ...


# ---------------------------------------------------------------------------
# Template generation
# ---------------------------------------------------------------------------


def build_ticket_template_workbook(
    *, catalog_repository: CatalogRepository, user_repository: UserRepository
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = TICKET_SHEET
    ws.append(list(TICKET_REQUIRED_HEADERS))

    services = catalog_repository.list(limit=500) or []
    incidents = [s for s in services if s.get("service_type") == "incident" and s.get("active", True)]
    requests = [s for s in services if s.get("service_type") == "service_request" and s.get("active", True)]

    inc_ref = wb.create_sheet(TICKET_REF_INCIDENT)
    inc_ref.append(["service_id", "name", "value_stream"])
    for s in incidents:
        inc_ref.append([s.get("service_id"), s.get("name"), s.get("value_stream")])

    req_ref = wb.create_sheet(TICKET_REF_SERVICE_REQUEST)
    req_ref.append(["service_id", "name", "value_stream"])
    for s in requests:
        req_ref.append([s.get("service_id"), s.get("name"), s.get("value_stream")])

    users_ref = wb.create_sheet(TICKET_REF_USERS)
    users_ref.append(["username", "display_name", "is_active"])
    # The user repository does not expose a list endpoint in this slice — the
    # template ships an empty reference sheet header. Frontend consumers can
    # resolve active users from the dedicated users endpoint.
    _populate_active_users(users_ref, user_repository)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _populate_active_users(ws, user_repository: UserRepository) -> None:
    """Best-effort user listing; failures fall back to header-only."""

    db = getattr(user_repository, "_session", None)
    list_active = getattr(user_repository, "list_active", None)
    if list_active is None:
        return
    try:
        rows = list_active(db) if db is not None else list_active()
    except Exception:  # noqa: BLE001 — never break template generation
        return
    for row in rows or []:
        ws.append([row.get("username"), row.get("display_name"), row.get("is_active", True)])


# ---------------------------------------------------------------------------
# Parsing + validation
# ---------------------------------------------------------------------------


def parse_ticket_workbook(
    payload: bytes,
    *,
    max_size_bytes: int = DEFAULT_MAX_SIZE_BYTES,
    catalog_repository: CatalogRepository | None = None,
    user_repository: UserRepository | None = None,
) -> list[dict]:
    guard_xlsx_payload(payload, max_size_bytes=max_size_bytes)
    wb = open_workbook(payload)

    headers = read_header_row(wb, TICKET_SHEET)
    header_error = validate_required_headers(headers, TICKET_REQUIRED_HEADERS)
    if header_error:
        raise header_error

    rows: list[dict] = []
    accumulated = ImportValidationError()
    for visible_row, cells in read_data_rows(wb, TICKET_SHEET):
        if not any(c.strip() for c in cells):
            continue
        normalized, row_error = _normalize_ticket_row(
            visible_row, cells, catalog_repository, user_repository
        )
        if row_error is not None:
            for err in row_error.errors:
                accumulated.add(
                    row=err.row,
                    field=err.field,
                    code=err.code,
                    reason=err.reason,
                )
            continue
        rows.append(normalized)

    if accumulated.has_errors():
        raise accumulated
    return rows


def _normalize_ticket_row(
    visible_row: int,
    cells: tuple[str, ...],
    catalog_repository: CatalogRepository | None,
    user_repository: UserRepository | None,
) -> tuple[dict | None, ImportValidationError | None]:
    error = ImportValidationError()

    def fail(field: str, code: str, reason: str) -> None:
        error.add(row=visible_row, field=field, code=code, reason=reason)

    raw_type = cells[0].strip() if len(cells) > 0 else ""
    raw_title = cells[1].strip() if len(cells) > 1 else ""
    raw_description = cells[2].strip() if len(cells) > 2 else ""
    raw_service = cells[3].strip() if len(cells) > 3 else ""
    raw_assignee = cells[4].strip() if len(cells) > 4 else ""

    if not raw_title:
        fail("title", "required", "title is required")
    if not raw_description:
        fail("description", "required", "description is required")
    if not raw_service:
        fail("service_catalog_id", "required", "service_catalog_id is required")
    if not raw_assignee:
        fail("assignee_username", "required", "assignee_username is required")

    if raw_type not in {"incident", "service_request"}:
        fail("type", "invalid_enum", "Must be one of: incident, service_request")

    if (
        catalog_repository is not None
        and raw_service
        and raw_type in {"incident", "service_request"}
    ):
        catalog = catalog_repository.get_by_id(raw_service)
        if catalog is None:
            fail("service_catalog_id", "service_not_found", f"Service '{raw_service}' does not exist")
        elif not catalog.get("active", True):
            fail("service_catalog_id", "service_inactive", f"Service '{raw_service}' is inactive")
        elif catalog.get("service_type") != raw_type:
            fail(
                "service_catalog_id",
                "service_type_mismatch",
                f"Service '{raw_service}' is type '{catalog.get('service_type')}', ticket is '{raw_type}'",
            )

    if user_repository is not None and raw_assignee:
        user_row = user_repository.get_by_username(None, raw_assignee)
        if user_row is None:
            fail("assignee_username", "user_not_found", f"User '{raw_assignee}' does not exist")
        elif not getattr(user_row, "is_active", True):
            fail("assignee_username", "user_inactive", f"User '{raw_assignee}' is inactive")

    if error.has_errors():
        return None, error

    return (
        {
            "type": raw_type,
            "title": raw_title,
            "description": raw_description,
            "service_catalog_id": raw_service,
            "assignee_username": raw_assignee,
        },
        None,
    )


# ---------------------------------------------------------------------------
# Atomic persistence with lock-ordered full-batch behavior
# ---------------------------------------------------------------------------


def import_ticket_workbook(
    payload: bytes,
    *,
    actor: str | None,
    ticket_repository: Any,
    catalog_repository: CatalogRepository | None = None,
    user_repository: UserRepository | None = None,
    pg_session: Any,
    max_size_bytes: int = DEFAULT_MAX_SIZE_BYTES,
) -> dict[str, Any]:
    rows = parse_ticket_workbook(
        payload,
        max_size_bytes=max_size_bytes,
        catalog_repository=catalog_repository,
        user_repository=user_repository,
    )

    # Final Pydantic validation against the canonical contract.
    normalized: list[TicketFolioCreate] = []
    error = ImportValidationError()
    for row in rows:
        try:
            normalized.append(TicketFolioCreate(**row))
        except ValidationError as exc:
            for issue in exc.errors():
                error.add(
                    row=None,
                    field=".".join(str(p) for p in issue["loc"]) or "workbook",
                    code="contract_violation",
                    reason=issue["msg"],
                )
    if error.has_errors():
        raise error

    # Lock-ordered full-batch acquisition — required by WU 7 to prevent
    # deadlock cycles when the same set of assignees is also being deactivated.
    if pg_session is not None and normalized:
        # Pre-sort/dedupe at the call site so the helper receives a stable
        # canonical order independent of the workbook's row order.
        ordered_assignees = sorted(
            {payload_model.assignee_username.lower() for payload_model in normalized}
        )
        try:
            acquire_user_locks_in_order(pg_session, ordered_assignees)
        except RuntimeError as exc:
            if "user_lock_timeout" in str(exc):
                raise ImportValidationError() from exc
            raise

    created = ticket_repository.bulk_create_with_generated_ids(normalized, actor=actor)
    return {
        "status": "imported",
        "imported_count": len(created),
        "rows": created,
    }


__all__ = [
    "TICKET_SHEET",
    "TICKET_REF_INCIDENT",
    "TICKET_REF_SERVICE_REQUEST",
    "TICKET_REF_USERS",
    "build_ticket_template_workbook",
    "import_ticket_workbook",
    "parse_ticket_workbook",
]
