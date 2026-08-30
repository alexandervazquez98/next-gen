"""Shared XLSX parsing helpers used by the catalog and ticket import services.

The helpers stay thin and return deterministic, machine-readable errors so
import routes can surface ``validation_failed`` payloads without touching the
workbook twice.
"""

from __future__ import annotations

from io import BytesIO
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .errors import ImportValidationError


DEFAULT_MAX_SIZE_BYTES = 5 * 1024 * 1024  # 5 MB


def guard_xlsx_payload(
    payload: bytes,
    *,
    max_size_bytes: int = DEFAULT_MAX_SIZE_BYTES,
) -> bytes:
    """Reject obviously bad payloads before openpyxl touches them."""

    if not isinstance(payload, (bytes, bytearray)) or len(payload) == 0:
        raise ImportValidationError() if False else _quick_error("workbook", "invalid_workbook", "Empty workbook payload")

    if len(payload) > max_size_bytes:
        raise _quick_error(
            "workbook",
            "oversized",
            f"Workbook exceeds maximum size of {max_size_bytes} bytes",
        )

    return bytes(payload)


def open_workbook(payload: bytes) -> Workbook:
    """Load a workbook from bytes; raise a structured error if openpyxl cannot read it."""

    try:
        return load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except (InvalidFileException, Exception):  # noqa: BLE001 — openpyxl raises a wide exception set
        raise _quick_error("workbook", "invalid_workbook", "Workbook is not a valid .xlsx file")


def read_header_row(workbook: Workbook, sheet_name: str) -> list[str]:
    if sheet_name not in workbook.sheetnames:
        raise _quick_error(
            sheet_name,
            "missing_sheet",
            f"Workbook is missing required sheet '{sheet_name}'",
        )
    ws = workbook[sheet_name]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    return [str(c).strip() if c is not None else "" for c in (header_row or [])]


def read_data_rows(
    workbook: Workbook, sheet_name: str
) -> Iterable[tuple[int, tuple[str, ...]]]:
    """Yield ``(visible_row_number, cell_tuple)`` for every row after the header."""

    if sheet_name not in workbook.sheetnames:
        raise _quick_error(
            sheet_name,
            "missing_sheet",
            f"Workbook is missing required sheet '{sheet_name}'",
        )
    ws = workbook[sheet_name]
    for visible_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cells = tuple("" if c is None else str(c) for c in row)
        yield visible_idx, cells


def validate_required_headers(
    actual: Sequence[str], required: Sequence[str]
) -> ImportValidationError | None:
    """Return an error container if any required header is missing."""

    actual_normalized = {h.strip() for h in actual}
    missing = [h for h in required if h not in actual_normalized]
    if not missing:
        return None
    err = ImportValidationError()
    for header in missing:
        err.add(row=None, field=header, code="missing_header", reason=f"Required header '{header}' is missing")
    return err


def reject_disallowed_headers(
    actual: Sequence[str], disallowed: Sequence[str]
) -> ImportValidationError | None:
    """Reject internal-field headers that should not appear in user-facing workbooks."""

    normalized = {h.strip() for h in actual}
    bad = [h for h in disallowed if h in normalized]
    if not bad:
        return None
    err = ImportValidationError()
    for header in bad:
        err.add(
            row=None,
            field=header,
            code="invalid_header",
            reason=f"Header '{header}' is not allowed; use 'SLA' instead",
        )
    return err


def _quick_error(field: str, code: str, reason: str) -> ImportValidationError:
    err = ImportValidationError()
    err.add(row=None, field=field, code=code, reason=reason)
    return err


__all__ = [
    "DEFAULT_MAX_SIZE_BYTES",
    "guard_xlsx_payload",
    "open_workbook",
    "read_data_rows",
    "read_header_row",
    "reject_disallowed_headers",
    "validate_required_headers",
]
