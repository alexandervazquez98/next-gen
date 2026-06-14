"""Service-level tests for node_service.py — mocks/stubs, no real DB.

Focus areas:
- get_nodes(): data scoping, metric parsing, SNMP deserialization, location handling
- create_update_node(): permission enforcement, ping metric creation, reconciliation triggers
- delete_node(): permission enforcement
- get_node_usage(): passthrough behavior
- bulk_upload_nodes(): Excel parsing, validation, status normalization, error aggregation
- get_node_template(): streaming response generation

Strategy:
- Patch topology_repo at the module level where it is imported (services.node_service.topology_repo)
- Patch check_permission from auth_service
- Patch reconcile_node_metrics from metric_service (lazy import inside create_update_node)
- Use conftest fixtures for User objects
"""

import pytest
import json
import io
import asyncio
from unittest.mock import patch, MagicMock, call
from datetime import datetime
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from fastapi.responses import JSONResponse

from models.user import User, UserRole, UserPermission
from models.core import Node


_UNSET = object()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_user(
    username: str = "testuser",
    role: str = "OPERATOR",
    permissions: list[UserPermission] | None = None,
    allowed_locations: list[str] | None = None,
    disabled: bool = False,
) -> User:
    return User(
        username=username,
        role=role,
        permissions=permissions or [],
        allowed_locations=allowed_locations or [],
        disabled=disabled,
    )


def _make_node_record(
    node_id: str = "ci-001",
    name: str = "Router-01",
    status: str = "OK",
    ip: str = "192.168.1.1",
    brand: str = "Cisco",
    model: str = "ASR-1000",
    layer: str = "router",
    location_name: str = "Data Center A",
    snmp: str | dict | None | object = _UNSET,
    extra_props: dict | None = None,
) -> dict:
    """Build a raw node dict as returned from Neo4j."""
    props = {
        "id": node_id,
        "name": name,
        "status": status,
        "ip": ip,
        "brand": brand,
        "model": model,
        "layer": layer,
        "location_name": location_name,
        "snmp": (
            json.dumps({"version": "v2c", "readCommunity": "public"})
            if snmp is _UNSET
            else snmp
        ),
        "location": None,
        "pollingInterval": 60,
    }
    if extra_props:
        props.update(extra_props)
    return props


def _make_full_record(
    node_props: dict | None = None,
    category: str = "router",
    metrics: list[dict] | None = None,
) -> dict:
    return {
        "node": node_props or _make_node_record(),
        "category": category,
        "metrics": metrics or [],
    }


def _run(coro):
    return asyncio.run(coro)


async def _collect_body(response: StreamingResponse) -> bytes:
    content = b""
    async for chunk in response.body_iterator:
        content += chunk
    return content


def _read_streaming_body(response: StreamingResponse) -> bytes:
    return asyncio.run(_collect_body(response))


# ---------------------------------------------------------------------------
# Tests: get_nodes()
# ---------------------------------------------------------------------------


class TestGetNodes:
    """Tests for the get_nodes service function."""

    def test_admin_gets_all_nodes_no_location_filter(self):
        """Admin user should get all nodes with is_admin=True."""
        admin = _make_user(username="admin", role="ADMIN")

        node_record = _make_full_record(
            node_props=_make_node_record(node_id="ci-001", name="Router-01"),
        )

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = [node_record]

            from services.node_service import get_nodes

            result = get_nodes(admin)

            assert len(result) == 1
            assert result[0]["label"] == "Router-01"
            mock_repo.get_nodes.assert_called_once()
            call_args = mock_repo.get_nodes.call_args
            assert call_args[0][0] == []  # allowed_locations
            assert call_args[0][1] is True  # is_admin

    def test_operator_scoped_by_location(self):
        """Operator should be scoped to allowed_locations."""
        operator = _make_user(
            username="op",
            role="OPERATOR",
            allowed_locations=["HQ-Madrid"],
        )

        node_record = _make_full_record(
            node_props=_make_node_record(node_id="ci-002", name="Switch-01"),
        )

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = [node_record]

            from services.node_service import get_nodes

            result = get_nodes(operator)

            assert len(result) == 1
            mock_repo.get_nodes.assert_called_once_with(["HQ-Madrid"], False)

    def test_operator_no_locations_returns_empty(self):
        """Operator with no allowed_locations should get empty list."""
        operator = _make_user(
            username="op",
            role="OPERATOR",
            allowed_locations=[],
        )

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = []

            from services.node_service import get_nodes

            result = get_nodes(operator)

            assert result == []
            mock_repo.get_nodes.assert_called_once_with([], False)

    def test_node_metrics_parsed_correctly(self):
        """Metrics from Neo4j records should be parsed into the response."""
        admin = _make_user(username="admin", role="ADMIN")
        ts = datetime(2026, 4, 1, 12, 0, 0)

        metrics = [
            {
                "name": "cpu-load",
                "protocol": "SNMP",
                "status": "OK",
                "value": 45.2,
                "last_updated": ts,
            },
            {
                "name": None,
                "protocol": "ICMP",
                "status": "UNKNOWN",
            },  # null name → filtered
        ]
        node_record = _make_full_record(
            node_props=_make_node_record(node_id="ci-001"),
            metrics=metrics,
        )

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = [node_record]

            from services.node_service import get_nodes

            result = get_nodes(admin)

            assert len(result[0]["metrics"]) == 1
            assert result[0]["metrics"][0]["name"] == "cpu-load"
            assert result[0]["metrics"][0]["value"] == 45.2
            assert result[0]["metrics"][0]["last_updated"] == ts.isoformat()

    def test_snmp_string_parsed_to_dict(self):
        """SNMP stored as JSON string should be deserialized."""
        admin = _make_user(username="admin", role="ADMIN")
        snmp_json = json.dumps({"version": "v3", "readCommunity": "secure"})
        node_record = _make_full_record(
            node_props=_make_node_record(snmp=snmp_json),
        )

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = [node_record]

            from services.node_service import get_nodes

            result = get_nodes(admin)

            assert isinstance(result[0]["snmp"], dict)
            assert result[0]["snmp"]["version"] == "v3"

    def test_snmp_none_stays_none(self):
        """SNMP that is None should remain None."""
        admin = _make_user(username="admin", role="ADMIN")
        node_record = _make_full_record(
            node_props=_make_node_record(snmp=None),
        )

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = [node_record]

            from services.node_service import get_nodes

            result = get_nodes(admin)

            assert result[0]["snmp"] is None

    def test_snmp_invalid_json_becomes_none(self):
        """SNMP that is an invalid JSON string should become None."""
        admin = _make_user(username="admin", role="ADMIN")
        node_record = _make_full_record(
            node_props=_make_node_record(snmp="not-valid-json{{{"),
        )

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = [node_record]

            from services.node_service import get_nodes

            result = get_nodes(admin)

            assert result[0]["snmp"] is None

    def test_location_object_serialized(self):
        """Location with latitude/longitude attributes should be serialized."""
        admin = _make_user(username="admin", role="ADMIN")

        class FakeLocation:
            def __init__(self):
                self.latitude = 19.4326
                self.longitude = -99.1332

        node_props = _make_node_record()
        node_props["location"] = FakeLocation()
        node_record = _make_full_record(node_props=node_props)

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = [node_record]

            from services.node_service import get_nodes

            result = get_nodes(admin)

            assert result[0]["location"] == {"lat": 19.4326, "long": -99.1332}

    def test_location_none_returns_none(self):
        """Node without location should have None location in output."""
        admin = _make_user(username="admin", role="ADMIN")
        node_record = _make_full_record(
            node_props=_make_node_record(),
        )

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = [node_record]

            from services.node_service import get_nodes

            result = get_nodes(admin)

            assert result[0]["location"] is None

    def test_metadata_excludes_standard_fields(self):
        """Metadata dict should exclude standard fields (id, name, status, etc.)."""
        admin = _make_user(username="admin", role="ADMIN")
        extra = {"custom_field": "value1", "another": 42}
        node_props = _make_node_record()
        node_props.update(extra)
        node_record = _make_full_record(node_props=node_props)

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = [node_record]

            from services.node_service import get_nodes

            result = get_nodes(admin)

            metadata = result[0]["metadata"]
            assert metadata["custom_field"] == "value1"
            assert metadata["another"] == 42
            # Standard fields should NOT be in metadata
            assert "id" not in metadata
            assert "name" not in metadata
            assert "status" not in metadata

    def test_type_defaults_to_layer_when_no_category(self):
        """When category is None, type should fall back to node layer."""
        admin = _make_user(username="admin", role="ADMIN")
        node_record = _make_full_record(
            node_props=_make_node_record(layer="firewall"),
            category=None,
        )

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = [node_record]

            from services.node_service import get_nodes

            result = get_nodes(admin)

            assert result[0]["type"] == "firewall"

    def test_type_uses_category_when_present(self):
        """When category exists, it should be used as type."""
        admin = _make_user(username="admin", role="ADMIN")
        node_record = _make_full_record(
            node_props=_make_node_record(layer="router"),
            category="core-router",
        )

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = [node_record]

            from services.node_service import get_nodes

            result = get_nodes(admin)

            assert result[0]["type"] == "core-router"

    def test_category_field_is_exposed_and_type_remains_compatible(self):
        """Nodes must include category metadata while preserving type semantics."""
        admin = _make_user(username="admin", role="ADMIN")
        node_record = _make_full_record(
            node_props=_make_node_record(layer="router"),
            category="Core Router",
        )

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = [
                {
                    **node_record,
                    "category_icon_key": "router",
                }
            ]

            from services.node_service import get_nodes

            result = get_nodes(admin)

            assert result[0]["type"] == "Core Router"
            assert result[0]["category"] == "Core Router"
            assert result[0]["category_icon_key"] == "router"

    def test_category_icon_key_defaults_to_generic_when_missing(self):
        """Missing icon metadata should default to generic in node payloads."""
        admin = _make_user(username="admin", role="ADMIN")
        node_record = _make_full_record(
            node_props=_make_node_record(layer="firewall"),
            category="Legacy Device",
            metrics=[],
        )

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = [
                {
                    **node_record,
                    "category_icon_key": None,
                }
            ]

            from services.node_service import get_nodes

            result = get_nodes(admin)

            assert result[0]["type"] == "Legacy Device"
            assert result[0]["category"] == "Legacy Device"
            assert result[0]["category_icon_key"] == "generic"

    def test_empty_nodes_list_returns_empty(self):
        """When repo returns no nodes, result should be empty list."""
        admin = _make_user(username="admin", role="ADMIN")

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = []

            from services.node_service import get_nodes

            result = get_nodes(admin)

            assert result == []

    def test_datetime_fields_serialized_to_iso(self):
        """Datetime fields in metadata should be serialized to ISO format."""
        admin = _make_user(username="admin", role="ADMIN")
        ts = datetime(2026, 3, 15, 10, 30, 0)
        node_props = _make_node_record()
        node_props["created_at"] = ts
        node_record = _make_full_record(node_props=node_props)

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_nodes.return_value = [node_record]

            from services.node_service import get_nodes

            result = get_nodes(admin)

            assert result[0]["metadata"]["created_at"] == ts.isoformat()


# ---------------------------------------------------------------------------
# Tests: create_update_node()
# ---------------------------------------------------------------------------


class TestCreateUpdateNode:
    """Tests for the create_update_node service function."""

    def test_denied_without_ci_edit_permission(self):
        """User without CI_EDIT should get 403."""
        viewer = _make_user(username="viewer", role="VIEWER", permissions=[])
        node = Node(id="ci-001", label="Router-01", type="router")

        with patch("services.node_service.check_permission", return_value=False):
            from services.node_service import create_update_node

            with pytest.raises(HTTPException) as exc_info:
                create_update_node(node, viewer)

            assert exc_info.value.status_code == 403
            assert "CI_EDIT" in exc_info.value.detail

    def test_admin_allowed_ci_edit(self):
        """Admin should pass permission check (check_permission returns True)."""
        admin = _make_user(username="admin", role="ADMIN")
        node = Node(id="ci-001", label="Router-01", type="router", ip="10.0.0.1")

        with (
            patch("services.node_service.check_permission", return_value=True),
            patch("services.node_service.topology_repo") as mock_repo,
            patch("services.metric_service.reconcile_node_metrics") as mock_reconcile,
        ):
            from services.node_service import create_update_node

            result = create_update_node(node, admin)

            assert result["message"] == "Node created/updated"
            assert result["id"] == "ci-001"
            mock_repo.upsert_node.assert_called_once_with(node)

    def test_operator_with_ci_edit_allowed(self):
        """Operator with CI_EDIT permission should succeed."""
        operator = _make_user(
            username="op",
            role="OPERATOR",
            permissions=[UserPermission.CI_EDIT],
        )
        node = Node(id="ci-002", label="Switch-01", type="switch")

        with (
            patch("services.node_service.check_permission", return_value=True),
            patch("services.node_service.topology_repo") as mock_repo,
            patch("services.metric_service.reconcile_node_metrics") as mock_reconcile,
        ):
            from services.node_service import create_update_node

            result = create_update_node(node, operator)

            assert result["message"] == "Node created/updated"
            mock_repo.upsert_node.assert_called_once()

    def test_creates_ping_metric_when_ip_present(self):
        """When node has an IP, a default PING metric should be created."""
        admin = _make_user(username="admin", role="ADMIN")
        node = Node(id="ci-001", label="Router-01", type="router", ip="192.168.1.1")

        with (
            patch("services.node_service.check_permission", return_value=True),
            patch("services.node_service.topology_repo") as mock_repo,
            patch("services.metric_service.reconcile_node_metrics"),
        ):
            from services.node_service import create_update_node

            create_update_node(node, admin)

            mock_repo.create_default_ping_metric.assert_called_once_with(
                "ci-001", "Router-01"
            )

    def test_no_ping_metric_when_ip_absent(self):
        """When node has no IP, no PING metric should be created."""
        admin = _make_user(username="admin", role="ADMIN")
        node = Node(id="ci-002", label="Server-01", type="server", ip=None)

        with (
            patch("services.node_service.check_permission", return_value=True),
            patch("services.node_service.topology_repo") as mock_repo,
            patch("services.metric_service.reconcile_node_metrics"),
        ):
            from services.node_service import create_update_node

            create_update_node(node, admin)

            mock_repo.create_default_ping_metric.assert_not_called()

    def test_reconciliation_called_on_create_update(self):
        """Metric reconciliation should be triggered after upsert."""
        admin = _make_user(username="admin", role="ADMIN")
        node = Node(
            id="ci-001",
            label="Router-01",
            type="router",
            brand="Cisco",
            model="ASR-1000",
        )

        with (
            patch("services.node_service.check_permission", return_value=True),
            patch("services.node_service.topology_repo"),
            patch("services.metric_service.reconcile_node_metrics") as mock_reconcile,
        ):
            from services.node_service import create_update_node

            create_update_node(node, admin)

            mock_reconcile.assert_called_once()
            # Should receive the node as a dict
            call_arg = mock_reconcile.call_args[0][0]
            assert isinstance(call_arg, dict)
            assert call_arg["id"] == "ci-001"

    def test_reconciliation_error_logged_but_does_not_fail(self):
        """If reconciliation fails, the operation should still succeed."""
        admin = _make_user(username="admin", role="ADMIN")
        node = Node(id="ci-001", label="Router-01", type="router")

        with (
            patch("services.node_service.check_permission", return_value=True),
            patch("services.node_service.topology_repo"),
            patch(
                "services.metric_service.reconcile_node_metrics",
                side_effect=Exception("DB timeout"),
            ),
        ):
            from services.node_service import create_update_node

            result = create_update_node(node, admin)

            # Should still return success
            assert result["message"] == "Node created/updated"
            assert result["id"] == "ci-001"

    def test_full_flow_order_upsert_then_ping_then_reconcile(self):
        """Verify the correct order: upsert → ping metric → reconcile."""
        admin = _make_user(username="admin", role="ADMIN")
        node = Node(id="ci-001", label="Router-01", type="router", ip="10.0.0.1")

        call_order = []

        def track_upsert(*args, **kwargs):
            call_order.append("upsert")

        def track_ping(*args, **kwargs):
            call_order.append("ping")

        def track_reconcile(*args, **kwargs):
            call_order.append("reconcile")

        with (
            patch("services.node_service.check_permission", return_value=True),
            patch("services.node_service.topology_repo") as mock_repo,
            patch("services.metric_service.reconcile_node_metrics") as mock_reconcile,
        ):
            mock_repo.upsert_node.side_effect = track_upsert
            mock_repo.create_default_ping_metric.side_effect = track_ping
            mock_reconcile.side_effect = track_reconcile

            from services.node_service import create_update_node

            create_update_node(node, admin)

            assert call_order == ["upsert", "ping", "reconcile"]


# ---------------------------------------------------------------------------
# Tests: delete_node()
# ---------------------------------------------------------------------------


class TestDeleteNode:
    """Tests for the delete_node service function."""

    def test_denied_without_ci_delete_permission(self):
        """User without CI_DELETE should get 403."""
        operator = _make_user(
            username="op",
            role="OPERATOR",
            permissions=[UserPermission.CI_EDIT],  # no CI_DELETE
        )

        with patch("services.node_service.check_permission", return_value=False):
            from services.node_service import delete_node

            with pytest.raises(HTTPException) as exc_info:
                delete_node("ci-001", operator)

            assert exc_info.value.status_code == 403
            assert "CI_DELETE" in exc_info.value.detail

    def test_admin_can_delete(self):
        """Admin should be able to delete a node."""
        admin = _make_user(username="admin", role="ADMIN")

        with (
            patch("services.node_service.check_permission", return_value=True),
            patch("services.node_service.topology_repo") as mock_repo,
        ):
            from services.node_service import delete_node

            result = delete_node("ci-001", admin)

            assert result["message"] == "Node deleted"
            assert result["id"] == "ci-001"
            mock_repo.delete_node.assert_called_once_with("ci-001")

    def test_operator_with_ci_delete_can_delete(self):
        """Operator with CI_DELETE permission should succeed."""
        operator = _make_user(
            username="op",
            role="OPERATOR",
            permissions=[UserPermission.CI_DELETE],
        )

        with (
            patch("services.node_service.check_permission", return_value=True),
            patch("services.node_service.topology_repo") as mock_repo,
        ):
            from services.node_service import delete_node

            result = delete_node("ci-002", operator)

            assert result["message"] == "Node deleted"
            mock_repo.delete_node.assert_called_once_with("ci-002")


# ---------------------------------------------------------------------------
# Tests: get_node_usage()
# ---------------------------------------------------------------------------


class TestGetNodeUsage:
    """Tests for the get_node_usage service function."""

    def test_returns_count_from_repo(self):
        """Should return the relationship count from the repo."""
        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_node_usage.return_value = 5

            from services.node_service import get_node_usage

            result = get_node_usage("ci-001")

            assert result == {"count": 5}
            mock_repo.get_node_usage.assert_called_once_with("ci-001")

    def test_returns_zero_when_no_relationships(self):
        """Should return 0 when node has no relationships."""
        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_node_usage.return_value = 0

            from services.node_service import get_node_usage

            result = get_node_usage("ci-999")

            assert result == {"count": 0}


# ---------------------------------------------------------------------------
# Tests: bulk_upload_nodes()
# ---------------------------------------------------------------------------


class TestBulkUploadNodes:
    """Tests for the bulk_upload_nodes service function."""

    def _make_fake_excel_bytes(self) -> bytes:
        """Create minimal valid xlsx bytes using openpyxl."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "ID",
                "Label",
                "NetworkLayer",
                "OperationalStatus",
                "Brand",
                "Model",
                "SerialNumber",
                "Firmware",
                "IP",
                "SNMP_Version",
                "SNMP_Read",
                "SNMP_Write",
                "Owner",
                "Location",
                "Latitude",
                "Longitude",
                "Criticality",
            ]
        )
        ws.append(
            [
                "CI-001",
                "Test Router",
                "router",
                "ACTIVE",
                "Cisco",
                "ASR-1000",
                "SN123",
                "17.3.1",
                "10.0.0.1",
                "v2c",
                "public",
                "private",
                "NetOps",
                "DC-A",
                19.4326,
                -99.1332,
                "High",
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_successful_bulk_upload(self):
        """Valid Excel file should process all rows."""
        excel_bytes = self._make_fake_excel_bytes()

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_valid_owners_and_layers.return_value = (
                {"NetOps"},
                {"router", "switch"},
            )
            mock_repo.bulk_insert_node = MagicMock()

            from services.node_service import bulk_upload_nodes

            result = _run(bulk_upload_nodes(excel_bytes, "import.xlsx"))

            assert mock_repo.bulk_insert_node.call_count == 1
            # Check the response
            content = result
            if isinstance(content, JSONResponse):
                data = json.loads(content.body)
            else:
                data = content
            assert "Successfully processed 1" in data["message"]

    def test_validation_error_invalid_owner(self):
        """Row with unknown owner should be skipped and reported."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "ID",
                "Label",
                "NetworkLayer",
                "OperationalStatus",
                "Brand",
                "Model",
                "SerialNumber",
                "Firmware",
                "IP",
                "SNMP_Version",
                "SNMP_Read",
                "SNMP_Write",
                "Owner",
                "Location",
                "Latitude",
                "Longitude",
                "Criticality",
            ]
        )
        ws.append(
            [
                "CI-001",
                "Bad Router",
                "router",
                "ACTIVE",
                "Cisco",
                "ASR-1000",
                "SN123",
                "17.3.1",
                "10.0.0.1",
                "v2c",
                "public",
                "private",
                "UnknownOwner",
                "DC-A",
                0.0,
                0.0,
                "Low",
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_valid_owners_and_layers.return_value = (
                {"NetOps"},
                {"router"},
            )
            mock_repo.bulk_insert_node = MagicMock()

            from services.node_service import bulk_upload_nodes

            result = _run(bulk_upload_nodes(buf.getvalue(), "import.xlsx"))

            assert mock_repo.bulk_insert_node.call_count == 0
            assert isinstance(result, JSONResponse)
            data = json.loads(result.body)
            assert (
                data["errors"][0]
                == "Row 2 (ID: CI-001): Owner 'UnknownOwner' not found."
            )

    def test_validation_error_invalid_layer(self):
        """Row with unknown NetworkLayer should be skipped and reported."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "ID",
                "Label",
                "NetworkLayer",
                "OperationalStatus",
                "Brand",
                "Model",
                "SerialNumber",
                "Firmware",
                "IP",
                "SNMP_Version",
                "SNMP_Read",
                "SNMP_Write",
                "Owner",
                "Location",
                "Latitude",
                "Longitude",
                "Criticality",
            ]
        )
        ws.append(
            [
                "CI-001",
                "Bad Router",
                "unknown-layer",
                "ACTIVE",
                "Cisco",
                "ASR-1000",
                "SN123",
                "17.3.1",
                "10.0.0.1",
                "v2c",
                "public",
                "private",
                "NetOps",
                "DC-A",
                0.0,
                0.0,
                "Low",
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_valid_owners_and_layers.return_value = (
                {"NetOps"},
                {"router"},
            )
            mock_repo.bulk_insert_node = MagicMock()

            from services.node_service import bulk_upload_nodes

            result = _run(bulk_upload_nodes(buf.getvalue(), "import.xlsx"))

            assert mock_repo.bulk_insert_node.call_count == 0
            assert isinstance(result, JSONResponse)
            data = json.loads(result.body)
            assert "NetworkLayer 'unknown-layer' not found" in data["errors"][0]

    def test_status_normalization_healthy_to_active(self):
        """Status 'HEALTHY' and 'OK' should be normalized to 'ACTIVE'."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "ID",
                "Label",
                "NetworkLayer",
                "OperationalStatus",
                "Brand",
                "Model",
                "SerialNumber",
                "Firmware",
                "IP",
                "SNMP_Version",
                "SNMP_Read",
                "SNMP_Write",
                "Owner",
                "Location",
                "Latitude",
                "Longitude",
                "Criticality",
            ]
        )
        ws.append(
            [
                "CI-001",
                "Router A",
                "router",
                "HEALTHY",
                "Cisco",
                "ASR-1000",
                "SN1",
                "1.0",
                "10.0.0.1",
                "v2c",
                "public",
                "private",
                "NetOps",
                "DC-A",
                0.0,
                0.0,
                "Low",
            ]
        )
        ws.append(
            [
                "CI-002",
                "Router B",
                "router",
                "OK",
                "Cisco",
                "ASR-1000",
                "SN2",
                "1.0",
                "10.0.0.2",
                "v2c",
                "public",
                "private",
                "NetOps",
                "DC-A",
                0.0,
                0.0,
                "Low",
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_valid_owners_and_layers.return_value = (
                {"NetOps"},
                {"router"},
            )
            mock_repo.bulk_insert_node = MagicMock()

            from services.node_service import bulk_upload_nodes

            _run(bulk_upload_nodes(buf.getvalue(), "import.xlsx"))

            # Both calls should have status='ACTIVE'
            calls = mock_repo.bulk_insert_node.call_args_list
            assert calls[0][0][3] == "ACTIVE"  # HEALTHY → ACTIVE
            assert calls[1][0][3] == "ACTIVE"  # OK → ACTIVE

    def test_status_normalization_warning_to_exception(self):
        """Status 'WARNING' and 'CRITICAL' should be normalized to 'EXCEPTION'."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "ID",
                "Label",
                "NetworkLayer",
                "OperationalStatus",
                "Brand",
                "Model",
                "SerialNumber",
                "Firmware",
                "IP",
                "SNMP_Version",
                "SNMP_Read",
                "SNMP_Write",
                "Owner",
                "Location",
                "Latitude",
                "Longitude",
                "Criticality",
            ]
        )
        ws.append(
            [
                "CI-001",
                "Router A",
                "router",
                "WARNING",
                "Cisco",
                "ASR-1000",
                "SN1",
                "1.0",
                "10.0.0.1",
                "v2c",
                "public",
                "private",
                "NetOps",
                "DC-A",
                0.0,
                0.0,
                "Low",
            ]
        )
        ws.append(
            [
                "CI-002",
                "Router B",
                "router",
                "CRITICAL",
                "Cisco",
                "ASR-1000",
                "SN2",
                "1.0",
                "10.0.0.2",
                "v2c",
                "public",
                "private",
                "NetOps",
                "DC-A",
                0.0,
                0.0,
                "Low",
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_valid_owners_and_layers.return_value = (
                {"NetOps"},
                {"router"},
            )
            mock_repo.bulk_insert_node = MagicMock()

            from services.node_service import bulk_upload_nodes

            _run(bulk_upload_nodes(buf.getvalue(), "import.xlsx"))

            calls = mock_repo.bulk_insert_node.call_args_list
            assert calls[0][0][3] == "EXCEPTION"  # WARNING → EXCEPTION
            assert calls[1][0][3] == "EXCEPTION"  # CRITICAL → EXCEPTION

    def test_status_normalization_maintenance(self):
        """Status 'MAINTENANCE' should remain as 'MAINTENANCE'."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "ID",
                "Label",
                "NetworkLayer",
                "OperationalStatus",
                "Brand",
                "Model",
                "SerialNumber",
                "Firmware",
                "IP",
                "SNMP_Version",
                "SNMP_Read",
                "SNMP_Write",
                "Owner",
                "Location",
                "Latitude",
                "Longitude",
                "Criticality",
            ]
        )
        ws.append(
            [
                "CI-001",
                "Router A",
                "router",
                "MAINTENANCE",
                "Cisco",
                "ASR-1000",
                "SN1",
                "1.0",
                "10.0.0.1",
                "v2c",
                "public",
                "private",
                "NetOps",
                "DC-A",
                0.0,
                0.0,
                "Low",
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_valid_owners_and_layers.return_value = (
                {"NetOps"},
                {"router"},
            )
            mock_repo.bulk_insert_node = MagicMock()

            from services.node_service import bulk_upload_nodes

            _run(bulk_upload_nodes(buf.getvalue(), "import.xlsx"))

            assert mock_repo.bulk_insert_node.call_args[0][3] == "MAINTENANCE"

    def test_status_normalization_unknown_defaults_to_active(self):
        """Unknown status should default to 'ACTIVE'."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "ID",
                "Label",
                "NetworkLayer",
                "OperationalStatus",
                "Brand",
                "Model",
                "SerialNumber",
                "Firmware",
                "IP",
                "SNMP_Version",
                "SNMP_Read",
                "SNMP_Write",
                "Owner",
                "Location",
                "Latitude",
                "Longitude",
                "Criticality",
            ]
        )
        ws.append(
            [
                "CI-001",
                "Router A",
                "router",
                "SOME_WEIRD_STATUS",
                "Cisco",
                "ASR-1000",
                "SN1",
                "1.0",
                "10.0.0.1",
                "v2c",
                "public",
                "private",
                "NetOps",
                "DC-A",
                0.0,
                0.0,
                "Low",
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_valid_owners_and_layers.return_value = (
                {"NetOps"},
                {"router"},
            )
            mock_repo.bulk_insert_node = MagicMock()

            from services.node_service import bulk_upload_nodes

            _run(bulk_upload_nodes(buf.getvalue(), "import.xlsx"))

            assert mock_repo.bulk_insert_node.call_args[0][3] == "ACTIVE"

    def test_auto_generated_id_when_empty(self):
        """When ID is empty, an auto-generated CI-XXXXXXXX ID should be used."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "ID",
                "Label",
                "NetworkLayer",
                "OperationalStatus",
                "Brand",
                "Model",
                "SerialNumber",
                "Firmware",
                "IP",
                "SNMP_Version",
                "SNMP_Read",
                "SNMP_Write",
                "Owner",
                "Location",
                "Latitude",
                "Longitude",
                "Criticality",
            ]
        )
        ws.append(
            [
                "",
                "Auto-ID Router",
                "router",
                "ACTIVE",
                "Cisco",
                "ASR-1000",
                "SN1",
                "1.0",
                "10.0.0.1",
                "v2c",
                "public",
                "private",
                "NetOps",
                "DC-A",
                0.0,
                0.0,
                "Low",
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_valid_owners_and_layers.return_value = (
                {"NetOps"},
                {"router"},
            )
            mock_repo.bulk_insert_node = MagicMock()

            from services.node_service import bulk_upload_nodes

            _run(bulk_upload_nodes(buf.getvalue(), "import.xlsx"))

            # The first positional arg is the generated ID
            generated_id = mock_repo.bulk_insert_node.call_args[0][0]
            assert generated_id.startswith("CI-")
            assert len(generated_id) == 11  # CI- + 8 hex chars

    def test_sample_row_skipped(self):
        """Header sample row ('Sample Router') should be skipped."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "ID",
                "Label",
                "NetworkLayer",
                "OperationalStatus",
                "Brand",
                "Model",
                "SerialNumber",
                "Firmware",
                "IP",
                "SNMP_Version",
                "SNMP_Read",
                "SNMP_Write",
                "Owner",
                "Location",
                "Latitude",
                "Longitude",
                "Criticality",
            ]
        )
        # This is the sample row from the template
        ws.append(
            [
                "(Leave Empty for Auto-ID)",
                "Sample Router",
                "INFRASTRUCTURE",
                "ACTIVE",
                "Cisco",
                "ASR-1000",
                "SN12345678",
                "17.3.1",
                "192.168.1.100",
                "v2c",
                "public",
                "private",
                "NetOps",
                "Data Center A",
                19.4326,
                -99.1332,
                "High",
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_valid_owners_and_layers.return_value = (
                {"NetOps"},
                {"router"},
            )
            mock_repo.bulk_insert_node = MagicMock()

            from services.node_service import bulk_upload_nodes

            _run(bulk_upload_nodes(buf.getvalue(), "import.xlsx"))

            mock_repo.bulk_insert_node.assert_not_called()

    def test_rows_with_empty_label_skipped(self):
        """Rows without a Label should be skipped."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "ID",
                "Label",
                "NetworkLayer",
                "OperationalStatus",
                "Brand",
                "Model",
                "SerialNumber",
                "Firmware",
                "IP",
                "SNMP_Version",
                "SNMP_Read",
                "SNMP_Write",
                "Owner",
                "Location",
                "Latitude",
                "Longitude",
                "Criticality",
            ]
        )
        ws.append(
            [
                "CI-001",
                "",
                "router",
                "ACTIVE",
                "Cisco",
                "ASR-1000",
                "SN1",
                "1.0",
                "10.0.0.1",
                "v2c",
                "public",
                "private",
                "NetOps",
                "DC-A",
                0.0,
                0.0,
                "Low",
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_valid_owners_and_layers.return_value = (
                {"NetOps"},
                {"router"},
            )
            mock_repo.bulk_insert_node = MagicMock()

            from services.node_service import bulk_upload_nodes

            _run(bulk_upload_nodes(buf.getvalue(), "import.xlsx"))

            mock_repo.bulk_insert_node.assert_not_called()

    def test_mixed_valid_and_invalid_rows_returns_207(self):
        """Mix of valid and invalid rows should return 207 with errors."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "ID",
                "Label",
                "NetworkLayer",
                "OperationalStatus",
                "Brand",
                "Model",
                "SerialNumber",
                "Firmware",
                "IP",
                "SNMP_Version",
                "SNMP_Read",
                "SNMP_Write",
                "Owner",
                "Location",
                "Latitude",
                "Longitude",
                "Criticality",
            ]
        )
        # Valid row
        ws.append(
            [
                "CI-001",
                "Good Router",
                "router",
                "ACTIVE",
                "Cisco",
                "ASR-1000",
                "SN1",
                "1.0",
                "10.0.0.1",
                "v2c",
                "public",
                "private",
                "NetOps",
                "DC-A",
                0.0,
                0.0,
                "Low",
            ]
        )
        # Invalid row (bad owner)
        ws.append(
            [
                "CI-002",
                "Bad Router",
                "router",
                "ACTIVE",
                "Cisco",
                "ASR-1000",
                "SN2",
                "1.0",
                "10.0.0.2",
                "v2c",
                "public",
                "private",
                "BadOwner",
                "DC-A",
                0.0,
                0.0,
                "Low",
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_valid_owners_and_layers.return_value = (
                {"NetOps"},
                {"router"},
            )
            mock_repo.bulk_insert_node = MagicMock()

            from services.node_service import bulk_upload_nodes

            result = _run(bulk_upload_nodes(buf.getvalue(), "import.xlsx"))

            assert isinstance(result, JSONResponse)
            assert result.status_code == 207
            data = json.loads(result.body)
            assert "Processed 1 CIs" in data["message"]
            assert len(data["errors"]) == 1

    def test_invalid_excel_raises_400(self):
        """Non-Excel content should raise HTTPException 400."""
        with pytest.raises(HTTPException) as exc_info:
            from services.node_service import bulk_upload_nodes

            _run(bulk_upload_nodes(b"not-an-excel-file", "import.xlsx"))

        assert exc_info.value.status_code == 400
        assert "Error reading Excel file" in exc_info.value.detail

    def test_multiple_valid_rows_processed(self):
        """Multiple valid rows should all be inserted."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "ID",
                "Label",
                "NetworkLayer",
                "OperationalStatus",
                "Brand",
                "Model",
                "SerialNumber",
                "Firmware",
                "IP",
                "SNMP_Version",
                "SNMP_Read",
                "SNMP_Write",
                "Owner",
                "Location",
                "Latitude",
                "Longitude",
                "Criticality",
            ]
        )
        for i in range(1, 4):
            ws.append(
                [
                    f"CI-{i:03d}",
                    f"Router-{i}",
                    "router",
                    "ACTIVE",
                    "Cisco",
                    "ASR-1000",
                    f"SN{i}",
                    "1.0",
                    f"10.0.0.{i}",
                    "v2c",
                    "public",
                    "private",
                    "NetOps",
                    "DC-A",
                    0.0,
                    0.0,
                    "Low",
                ]
            )
        buf = io.BytesIO()
        wb.save(buf)

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_valid_owners_and_layers.return_value = (
                {"NetOps"},
                {"router"},
            )
            mock_repo.bulk_insert_node = MagicMock()

            from services.node_service import bulk_upload_nodes

            result = _run(bulk_upload_nodes(buf.getvalue(), "import.xlsx"))

            assert mock_repo.bulk_insert_node.call_count == 3
            content = result
            if isinstance(content, JSONResponse):
                data = json.loads(content.body)
            else:
                data = content
            assert "Successfully processed 3" in data["message"]


# ---------------------------------------------------------------------------
# Tests: get_node_template()
# ---------------------------------------------------------------------------


class TestGetNodeTemplate:
    """Tests for the get_node_template service function."""

    def test_returns_streaming_response(self):
        """Should return a StreamingResponse with Excel content."""
        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_template_data.return_value = (
                ["NetOps", "SecOps"],
                ["router", "switch"],
            )

            from services.node_service import get_node_template

            result = get_node_template()

            assert isinstance(result, StreamingResponse)

    def test_template_contains_expected_sheets(self):
        """Template should contain Data Entry, Owners, and Layers sheets."""
        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_template_data.return_value = (["NetOps"], ["router"])

            from services.node_service import get_node_template

            result = get_node_template()

            # Read the streaming content
            content = _read_streaming_body(result)

            import openpyxl

            wb = openpyxl.load_workbook(filename=io.BytesIO(content))
            sheet_names = wb.sheetnames
            assert "Data Entry Template" in sheet_names
            assert "Ref - Owners" in sheet_names
            assert "Ref - Network Layers" in sheet_names

    def test_template_owners_populated_from_repo(self):
        """Owners reference sheet should contain data from the repo."""
        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_template_data.return_value = (
                ["NetOps", "SecOps", "CloudOps"],
                ["router"],
            )

            from services.node_service import get_node_template

            result = get_node_template()

            content = _read_streaming_body(result)

            import openpyxl
            import pandas as pd

            owners_df = pd.read_excel(io.BytesIO(content), sheet_name="Ref - Owners")
            assert list(owners_df["Available Owners"]) == [
                "NetOps",
                "SecOps",
                "CloudOps",
            ]

    def test_template_layers_populated_from_repo(self):
        """Layers reference sheet should contain data from the repo."""
        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_template_data.return_value = (
                ["NetOps"],
                ["router", "switch", "firewall"],
            )

            from services.node_service import get_node_template

            result = get_node_template()

            content = _read_streaming_body(result)

            import pandas as pd

            layers_df = pd.read_excel(
                io.BytesIO(content), sheet_name="Ref - Network Layers"
            )
            assert list(layers_df["Available Network Layers"]) == [
                "router",
                "switch",
                "firewall",
            ]

    def test_template_has_sample_data(self):
        """Data Entry Template should contain the sample router row."""
        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.get_template_data.return_value = ([], [])

            from services.node_service import get_node_template

            result = get_node_template()

            content = _read_streaming_body(result)

            import pandas as pd

            df = pd.read_excel(io.BytesIO(content), sheet_name="Data Entry Template")
            assert len(df) == 1
            assert df.iloc[0]["Label"] == "Sample Router"
            assert df.iloc[0]["Brand"] == "Cisco"


# ---------------------------------------------------------------------------
# Tests: search_nodes()
# ---------------------------------------------------------------------------


class TestSearchNodes:
    """Tests for the search_nodes service function."""

    def test_admin_search_returns_all_nodes_no_location_filter(self):
        """Admin user should search all nodes with is_admin=True, no location filter."""
        admin = _make_user(username="admin", role="ADMIN")

        node_record = _make_full_record(
            node_props=_make_node_record(node_id="ci-001", name="Router-01"),
        )

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.search_nodes.return_value = [
                {
                    "id": "ci-001",
                    "label": "Router-01",
                    "ip": "192.168.1.1",
                    "status": "OK",
                    "brand": "Cisco",
                    "model": "ASR-1000",
                }
            ]

            from services.node_service import search_nodes

            result = search_nodes(admin, "Router")

            assert len(result) == 1
            assert result[0]["label"] == "Router-01"
            mock_repo.search_nodes.assert_called_once()
            call_args = mock_repo.search_nodes.call_args
            # Admin: is_admin=True (arg 2), term passed as keyword arg
            assert call_args[0][0] == "Router"  # term
            assert call_args[0][2] is True  # is_admin

    def test_operator_search_scoped_by_location(self):
        """Operator should have results scoped to allowed_locations."""
        operator = _make_user(
            username="op",
            role="OPERATOR",
            allowed_locations=["HQ-Madrid"],
        )

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.search_nodes.return_value = [
                {
                    "id": "ci-002",
                    "label": "Switch-01",
                    "ip": "192.168.1.2",
                    "status": "OK",
                    "brand": "Cisco",
                    "model": "Catalyst",
                }
            ]

            from services.node_service import search_nodes

            result = search_nodes(operator, "Switch")

            assert len(result) == 1
            mock_repo.search_nodes.assert_called_once()
            call_args = mock_repo.search_nodes.call_args
            # Non-admin: passes term, allowed_locations, is_admin=False
            assert call_args[0][0] == "Switch"  # term
            assert call_args[0][1] == ["HQ-Madrid"]  # allowed_locations
            assert call_args[0][2] is False  # is_admin

    def test_operator_no_locations_returns_empty(self):
        """Operator with no allowed_locations should return empty list."""
        operator = _make_user(
            username="op",
            role="OPERATOR",
            allowed_locations=[],
        )

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.search_nodes.return_value = []

            from services.node_service import search_nodes

            result = search_nodes(operator, "Router")

            assert result == []
            mock_repo.search_nodes.assert_called_once()
            call_args = mock_repo.search_nodes.call_args
            # Empty list passed for non-admin with no scopes: term='Router', allowed_locations=[], is_admin=False
            assert call_args[0][0] == "Router"  # term
            assert call_args[0][1] == []  # allowed_locations

    def test_search_passes_term_correctly(self):
        """Search term should be passed to repo unchanged."""
        admin = _make_user(username="admin", role="ADMIN")

        with patch("services.node_service.topology_repo") as mock_repo:
            mock_repo.search_nodes.return_value = []

            from services.node_service import search_nodes

            search_nodes(admin, "Madrid")

            call_args = mock_repo.search_nodes.call_args
            assert call_args[0][0] == "Madrid"  # term
