"""Atomic XLSX catalog and ticket import — WU 6 + WU 7 (RED).

TDD failing-first coverage for the Service Management workbook import pipeline.
Tests are kept lean so the production implementation can fit inside the
800-line PR review budget.
"""

from __future__ import annotations

from collections.abc import Iterable
from io import BytesIO
from unittest.mock import MagicMock

import pytest
from openpyxl import Workbook, load_workbook
from services.itsm_imports import catalog_import, ticket_import
from services.itsm_imports.errors import (
    IMPORT_ERROR_CAP,
    IMPORT_VALIDATION_FAILED,
    ImportValidationError,
)

# ---------------------------------------------------------------------------
# Shared stubs
# ---------------------------------------------------------------------------


class _StubValueStreamLookup:
    def __init__(self, active: set[str] | None = None) -> None:
        self._active = active or {"operate", "deliver"}

    def list_active(self) -> list[dict]:
        return [{"value": v, "label": v.title()} for v in sorted(self._active)]

    def is_active(self, value: str) -> bool:
        return value in self._active


class _StubCatalogRepository:
    def __init__(self, services: dict[str, dict] | None = None) -> None:
        self._services = services or {}
        self.get_by_id = MagicMock(side_effect=self._get)
        self.list = MagicMock(return_value=list(self._services.values()))

    def _get(self, service_id: str):
        return self._services.get(service_id)


class _StubUserRepository:
    def __init__(self, users: Iterable[str], *, active: Iterable[str] | None = None) -> None:
        active_set = set(active if active is not None else users)
        self._rows = {name: MagicMock(username=name, is_active=(name in active_set)) for name in users}

    def get_by_username(self, _db, username: str):
        return self._rows.get(username)


# ---------------------------------------------------------------------------
# Workbook builders
# ---------------------------------------------------------------------------


def _build_catalog_workbook(
    rows: list[list[str]] | None = None,
    *,
    headers: list[str] | None = None,
    active_streams: set[str] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalog Import"
    ws.append(headers or ["service_id", "name", "SLA", "description", "service_type", "value_stream", "active"])
    for row in rows or []:
        ws.append(row)
    streams = active_streams if active_streams is not None else {"operate", "deliver"}
    ref = wb.create_sheet("Ref - Value Streams")
    ref.append(["value", "label"])
    for s in sorted(streams):
        ref.append([s, s.title()])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_ticket_workbook(rows: list[list[str]] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ticket Import"
    ws.append(["type", "title", "description", "service_catalog_id", "assignee_username"])
    for row in rows or []:
        ws.append(row)
    for name, ref_rows in (
        ("Ref - Incident Services", [["service_id", "name", "value_stream"], ["svc-inc-1", "Net", "operate"]]),
        ("Ref - Service Request Services", [["service_id", "name", "value_stream"], ["svc-req-1", "Access", "deliver"]]),
        ("Ref - Active Users", [["username", "display_name", "is_active"], ["op1", "Op One", "true"]]),
    ):
        ref = wb.create_sheet(name)
        for row in ref_rows:
            ref.append(row)
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _make_catalog_repo(rows: list[list[str]]) -> _StubCatalogRepository:
    return _StubCatalogRepository(
        {
            "svc-inc-1": {
                "service_id": "svc-inc-1",
                "service_type": "incident",
                "active": True,
                "value_stream": "operate",
            },
            "svc-req-1": {
                "service_id": "svc-req-1",
                "service_type": "service_request",
                "active": True,
                "value_stream": "deliver",
            },
        }
    )


# ---------------------------------------------------------------------------
# WU 6 — catalog import
# ---------------------------------------------------------------------------


class TestCatalogTemplate:
    def test_template_emits_required_sheets(self):
        wb_bytes = catalog_import.build_catalog_template_workbook(
            value_stream_lookup=_StubValueStreamLookup()
        )
        wb = load_workbook(BytesIO(wb_bytes))
        assert {"Catalog Import", "Ref - Value Streams"}.issubset(set(wb.sheetnames))
        headers = [c.value for c in wb["Catalog Import"][1]]
        assert "SLA" in headers
        assert "sla_target_minutes" not in headers


class TestCatalogHeaderValidation:
    def test_missing_required_header_returns_structured_error(self):
        bad = _build_catalog_workbook(
            headers=["service_id", "name", "SLA", "service_type", "value_stream", "active"]
        )
        with pytest.raises(ImportValidationError) as exc:
            catalog_import.parse_catalog_workbook(bad)
        assert exc.value.status == IMPORT_VALIDATION_FAILED
        fields = {e.field for e in exc.value.errors}
        assert "description" in fields

    def test_sla_target_minutes_header_is_rejected(self):
        bad = _build_catalog_workbook(
            headers=[
                "service_id",
                "name",
                "sla_target_minutes",
                "description",
                "service_type",
                "value_stream",
                "active",
            ]
        )
        with pytest.raises(ImportValidationError):
            catalog_import.parse_catalog_workbook(bad)


class TestCatalogRowValidation:
    def test_invalid_service_type_returns_invalid_enum(self):
        wb_bytes = _build_catalog_workbook([["svc-1", "Net", "30", "Desc", "request", "operate", "true"]])
        with pytest.raises(ImportValidationError) as exc:
            catalog_import.parse_catalog_workbook(wb_bytes)
        assert any(e.code == "invalid_enum" for e in exc.value.errors)

    def test_inactive_value_stream_rejected(self):
        wb_bytes = _build_catalog_workbook([["svc-1", "Net", "30", "Desc", "incident", "legacy", "true"]])
        with pytest.raises(ImportValidationError) as exc:
            catalog_import.parse_catalog_workbook(
                wb_bytes, value_stream_lookup=_StubValueStreamLookup()
            )
        assert any(e.code == "inactive_value_stream" for e in exc.value.errors)

    def test_error_payload_caps_results(self):
        rows = [["", "", "-1", "", "request", "legacy", "true"]] * (IMPORT_ERROR_CAP + 5)
        wb_bytes = _build_catalog_workbook(rows)
        with pytest.raises(ImportValidationError) as exc:
            catalog_import.parse_catalog_workbook(
                wb_bytes, value_stream_lookup=_StubValueStreamLookup()
            )
        assert len(exc.value.errors) == IMPORT_ERROR_CAP


class TestCatalogAtomicity:
    def test_invalid_workbook_persists_zero_rows(self):
        repository = MagicMock()
        repository.bulk_create = MagicMock(return_value=[])
        wb_bytes = _build_catalog_workbook(
            [
                ["svc-1", "Net", "30", "Desc", "incident", "operate", "true"],
                ["svc-2", "Bad", "-1", "Desc", "request", "operate", "true"],
            ]
        )
        with pytest.raises(ImportValidationError):
            catalog_import.import_catalog_workbook(
                wb_bytes,
                actor="admin",
                repository=repository,
                value_stream_lookup=_StubValueStreamLookup(),
            )
        repository.bulk_create.assert_not_called()

    def test_valid_workbook_persists_all_rows(self):
        repository = MagicMock()
        repository.bulk_create = MagicMock(
            return_value=[{"service_id": "svc-1"}, {"service_id": "svc-2"}]
        )
        wb_bytes = _build_catalog_workbook(
            [
                ["svc-1", "Net", "30", "Desc", "incident", "operate", "true"],
                ["svc-2", "Access", "60", "Grant", "service_request", "deliver", "true"],
            ]
        )
        result = catalog_import.import_catalog_workbook(
            wb_bytes,
            actor="admin",
            repository=repository,
            value_stream_lookup=_StubValueStreamLookup(),
        )
        assert result["status"] == "imported"
        assert result["imported_count"] == 2
        persisted = repository.bulk_create.call_args.args[0]
        assert all("sla_target_minutes" in row for row in persisted)
        assert all("SLA" not in row for row in persisted)


class TestCatalogFileGuard:
    def test_non_xlsx_payload_rejected(self):
        with pytest.raises(ImportValidationError) as exc:
            catalog_import.parse_catalog_workbook(b"plain text payload")
        assert any(e.code == "invalid_workbook" for e in exc.value.errors)


# ---------------------------------------------------------------------------
# WU 7 — ticket import
# ---------------------------------------------------------------------------


class TestTicketTemplate:
    def test_template_includes_three_reference_sheets(self):
        catalog_repo = _StubCatalogRepository(
            {"svc-inc-1": {"service_id": "svc-inc-1", "service_type": "incident", "active": True, "value_stream": "operate"}}
        )
        user_repo = _StubUserRepository(["op1"])
        wb_bytes = ticket_import.build_ticket_template_workbook(
            catalog_repository=catalog_repo,
            user_repository=user_repo,
        )
        wb = load_workbook(BytesIO(wb_bytes))
        expected = {
            "Ticket Import",
            "Ref - Incident Services",
            "Ref - Service Request Services",
            "Ref - Active Users",
        }
        assert expected.issubset(set(wb.sheetnames))


class TestTicketRowValidation:
    def test_missing_assignee_reports_row_error(self):
        wb_bytes = _build_ticket_workbook(
            [["incident", "Router down", "desc", "svc-inc-1", ""]]
        )
        with pytest.raises(ImportValidationError) as exc:
            ticket_import.parse_ticket_workbook(wb_bytes)
        assert any(e.field == "assignee_username" for e in exc.value.errors)

    def test_incompatible_service_type_reports_row_error(self):
        wb_bytes = _build_ticket_workbook(
            [["incident", "Router down", "desc", "svc-req-1", "op1"]]
        )
        with pytest.raises(ImportValidationError) as exc:
            ticket_import.parse_ticket_workbook(
                wb_bytes,
                catalog_repository=_StubCatalogRepository(
                    {"svc-req-1": {"service_id": "svc-req-1", "service_type": "service_request", "active": True, "value_stream": "deliver"}}
                ),
                user_repository=_StubUserRepository(["op1"]),
            )
        assert any(e.code == "service_type_mismatch" for e in exc.value.errors)

    def test_inactive_user_reports_row_error(self):
        wb_bytes = _build_ticket_workbook(
            [["incident", "Router down", "desc", "svc-inc-1", "op1"]]
        )
        catalog_repo = _make_catalog_repo([])
        with pytest.raises(ImportValidationError) as exc:
            ticket_import.parse_ticket_workbook(
                wb_bytes,
                catalog_repository=catalog_repo,
                user_repository=_StubUserRepository(["op1"], active=[]),
            )
        assert any(e.field == "assignee_username" for e in exc.value.errors)


class TestTicketAtomicityAndLocking:
    def test_invalid_workbook_persists_zero_tickets(self):
        wb_bytes = _build_ticket_workbook(
            [["incident", "Router down", "desc", "svc-inc-1", ""]]
        )
        ticket_repo = MagicMock()
        ticket_repo.bulk_create_with_generated_ids = MagicMock(return_value=[])
        with pytest.raises(ImportValidationError):
            ticket_import.import_ticket_workbook(
                wb_bytes,
                actor="admin",
                ticket_repository=ticket_repo,
                catalog_repository=_make_catalog_repo([]),
                user_repository=_StubUserRepository(["op1"]),
                pg_session=MagicMock(),
            )
        ticket_repo.bulk_create_with_generated_ids.assert_not_called()

    def test_valid_workbook_persists_all_tickets(self):
        wb_bytes = _build_ticket_workbook(
            [
                ["incident", "Router down", "desc", "svc-inc-1", "op1"],
                ["service_request", "Access", "grant", "svc-req-1", "op1"],
            ]
        )
        ticket_repo = MagicMock()
        ticket_repo.bulk_create_with_generated_ids = MagicMock(
            return_value=[{"ticket_id": 1}, {"ticket_id": 2}]
        )
        result = ticket_import.import_ticket_workbook(
            wb_bytes,
            actor="admin",
            ticket_repository=ticket_repo,
            catalog_repository=_make_catalog_repo([]),
            user_repository=_StubUserRepository(["op1"]),
            pg_session=MagicMock(),
        )
        assert result["status"] == "imported"
        assert result["imported_count"] == 2
        ticket_repo.bulk_create_with_generated_ids.assert_called_once()

    def test_lock_acquisition_is_sorted_and_deduped(self):
        wb_bytes = _build_ticket_workbook(
            [
                ["incident", "Router down", "desc", "svc-inc-1", "op2"],
                ["incident", "Backup", "desc", "svc-inc-1", "op1"],
            ]
        )
        ticket_repo = MagicMock()
        ticket_repo.bulk_create_with_generated_ids = MagicMock(
            return_value=[{"ticket_id": 1}, {"ticket_id": 2}]
        )
        captured: dict = {}

        def fake_locks(session, usernames):
            captured["usernames"] = list(usernames)
            return sorted({u.lower() for u in usernames})

        with pytest.MonkeyPatch.context() as mp:
            mp.setattr(
                "services.itsm_imports.ticket_import.acquire_user_locks_in_order",
                fake_locks,
            )
            ticket_import.import_ticket_workbook(
                wb_bytes,
                actor="admin",
                ticket_repository=ticket_repo,
                catalog_repository=_make_catalog_repo([]),
                user_repository=_StubUserRepository(["op1", "op2"]),
                pg_session=MagicMock(),
            )
        assert captured["usernames"] == ["op1", "op2"]
